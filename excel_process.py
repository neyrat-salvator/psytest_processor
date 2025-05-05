import os
import openpyxl
from openpyxl.styles import Font, Color, Alignment, Border, Side


# Метод для формирования главных атрибутов Excel-файла
def get_workbook(filename: str):
    
    file = ''.join([os.path.dirname(os.path.abspath(__file__)), '\\', filename])
    workbook = openpyxl.load_workbook(filename=file)
    
    return workbook


# Метод для выбора листа
def get_sheet_data(workbook, sheet_name: str):
    
    sheet = workbook.active
    sheet = workbook[sheet_name]
    sheet_data = {'sheet': sheet, 
                  'columns': list(sheet.column_dimensions)}
    
    return sheet_data
    

# Метод для установки значения в ячейку
def set_value_to_cells(sheet, target_column: str, row_index: int, value):
    
    sheet[f'{target_column}{row_index}'].value = str(value)
    # sheet.cell(column=target_column, row=row_index, value=value)


# Метод для считывания данных с ячейки
def get_value_from_cell(sheet, target_column: int, row_index: int):
    
    return sheet.cell(column=target_column, row=row_index).value


# Метод для сохранения и закрытия файла
def save_and_close(filename: str, workbook):
    
    filepath = os.path.dirname(os.path.abspath(__file__))
    new_filename = filename.rstrip('.xlsx') + '_Done.xlsx'
    new_filepath_name = ''.join([filepath, '\\', new_filename])
    workbook.save(new_filepath_name)
    workbook.close